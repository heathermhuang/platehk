#!/usr/bin/env python3
from __future__ import annotations

import json
import re
import shutil
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from urllib.parse import quote

import xlrd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
OUT = DATA / "tvrm_legacy"
ISSUES_DIR = OUT / "issues"

LEGACY_XLS_NAME = "TVRM auction result (1973-2026).xls"
LEGACY_XLS_PATH = DATA / LEGACY_XLS_NAME
EXACT_XLSX_NAME = "TVRM auction result (2006-2026).xlsx"
EXACT_XLSX_PATH = DATA / EXACT_XLSX_NAME

LEGACY_SOURCE_URL = f"./data/{quote(LEGACY_XLS_NAME)}"
LEGACY_END_YEAR = 2006
YEAR_RANGE_RE = re.compile(r"^(\d{4})-(\d{4})$")
PLATE_RE = re.compile(r"^([A-Z]{1,2})(\d{1,4})$")


def issue_key_for_range(year_range: str) -> str:
    match = YEAR_RANGE_RE.fullmatch(year_range)
    if not match:
        raise ValueError(f"invalid year range: {year_range}")
    return f"{match.group(1)}-01-01"


def normalize_plate(raw: str) -> str:
    value = re.sub(r"\s+", "", str(raw or "").upper())
    match = PLATE_RE.fullmatch(value)
    if match:
        return f"{match.group(1)} {match.group(2)}"
    return value


def normalize_amount(raw) -> int | None:
    text = str(raw or "").strip().replace(",", "")
    if text == "":
        return None
    try:
        return int(round(float(text)))
    except (TypeError, ValueError):
        return None


def normalize_date(raw) -> str | None:
    if isinstance(raw, datetime):
        return raw.date().isoformat()
    if isinstance(raw, date):
        return raw.isoformat()
    text = str(raw or "").strip()
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        return text
    return None


def result_status(text: str) -> str:
    text = str(text or "").strip()
    if "特別費用分配" in text:
        return "assigned_special_fee"
    if "售出" in text:
        return "sold"
    return "unknown"


def overlap_signature(plate: str, amount: int | None, status: str) -> str:
    return json.dumps([plate, amount, status], ensure_ascii=False, separators=(",", ":"))


def compare_row(row: dict) -> tuple:
    amount = row.get("amount_hkd")
    return (
        -(amount if amount is not None else -1),
        str(row.get("auction_date") or ""),
        str(row.get("single_line") or ""),
    )


def dedupe_key(row: dict) -> tuple:
    return (
        str(row.get("single_line") or ""),
        row.get("amount_hkd"),
        str(row.get("result_status") or ""),
    )


def dedupe_issue_rows(rows: list[dict]) -> list[dict]:
    deduped = {dedupe_key(row): dict(row) for row in rows}
    out = list(deduped.values())
    out.sort(key=compare_row)
    return out


def load_exact_overlap_signatures() -> set[str]:
    signatures: set[str] = set()
    wb = load_workbook(EXACT_XLSX_PATH, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                plate_raw = row[0] if len(row) > 0 else None
                auction_raw = row[1] if len(row) > 1 else None
                amount_raw = row[2] if len(row) > 2 else None
                result_raw = row[3] if len(row) > 3 else None

                plate = normalize_plate(plate_raw)
                if not plate:
                    continue
                auction_date = normalize_date(auction_raw)
                if not auction_date or int(auction_date[:4]) <= LEGACY_END_YEAR:
                    continue
                amount = normalize_amount(amount_raw)
                status = result_status(result_raw or "")
                signatures.add(overlap_signature(plate, amount, status))
    finally:
        wb.close()
    return signatures


def remapped_legacy_range(sheet_name: str) -> str | None:
    match = YEAR_RANGE_RE.fullmatch(sheet_name)
    if not match:
        return None
    start_year, end_year = (int(match.group(1)), int(match.group(2)))
    if start_year > LEGACY_END_YEAR:
        return None
    if end_year > LEGACY_END_YEAR:
        end_year = LEGACY_END_YEAR
    return f"{start_year}-{end_year}"


def load_legacy_rows(exact_signatures: set[str]) -> tuple[dict[str, list[dict]], dict[str, int]]:
    by_issue: dict[str, list[dict]] = defaultdict(list)
    stats = {
        "legacy_rows_skipped_due_to_exact_dates_after_2006": 0,
        "legacy_rows_kept_after_cutoff": 0,
        "legacy_source_sheet_count": 0,
    }

    book = xlrd.open_workbook(str(LEGACY_XLS_PATH))
    for sheet_name in book.sheet_names():
        remapped_range = remapped_legacy_range(sheet_name)
        if not remapped_range:
            continue
        stats["legacy_source_sheet_count"] += 1
        issue_key = issue_key_for_range(remapped_range)
        sheet = book.sheet_by_name(sheet_name)
        for r in range(1, sheet.nrows):
            plate = normalize_plate(sheet.cell_value(r, 0))
            if not plate:
                continue
            amount = normalize_amount(sheet.cell_value(r, 1))
            result_text = str(sheet.cell_value(r, 2) or "").strip()
            status = result_status(result_text)
            if overlap_signature(plate, amount, status) in exact_signatures:
                stats["legacy_rows_skipped_due_to_exact_dates_after_2006"] += 1
                continue
            stats["legacy_rows_kept_after_cutoff"] += 1
            by_issue[issue_key].append(
                {
                    "auction_date": issue_key,
                    "auction_date_label": remapped_range,
                    "date_precision": "year_range",
                    "year_range": remapped_range,
                    "single_line": plate,
                    "double_line": None,
                    "amount_hkd": amount,
                    "pdf_url": LEGACY_SOURCE_URL,
                    "source_url": LEGACY_SOURCE_URL,
                    "source_format": "xls",
                    "source_type": "xls_legacy_year_range",
                    "source_sheet": sheet_name,
                    "result_status": status,
                    "result_text": result_text,
                }
            )
    return by_issue, stats


def build() -> int:
    if not LEGACY_XLS_PATH.exists():
        raise FileNotFoundError(LEGACY_XLS_PATH)
    if not EXACT_XLSX_PATH.exists():
        raise FileNotFoundError(EXACT_XLSX_PATH)

    OUT.mkdir(parents=True, exist_ok=True)
    if ISSUES_DIR.exists():
        shutil.rmtree(ISSUES_DIR)
    ISSUES_DIR.mkdir(parents=True, exist_ok=True)

    exact_signatures = load_exact_overlap_signatures()
    by_issue, stats = load_legacy_rows(exact_signatures)

    all_rows: list[dict] = []
    auctions: list[dict] = []
    manifest_items: list[dict] = []

    issue_dates_desc = sorted(by_issue.keys(), reverse=True)
    for issue_key in issue_dates_desc:
        rows = dedupe_issue_rows(by_issue[issue_key])
        if not rows:
            continue
        label = str(rows[0].get("auction_date_label") or issue_key)
        year_range = rows[0].get("year_range")
        total_amount = sum(int(row["amount_hkd"]) for row in rows if row.get("amount_hkd") is not None)

        all_rows.extend(rows)
        shard_rel = f"issues/{issue_key}.json"
        (ISSUES_DIR / f"{issue_key}.json").write_text(
            json.dumps(rows, ensure_ascii=False, separators=(",", ":")),
            encoding="utf-8",
        )
        manifest_items.append(
            {
                "auction_date": issue_key,
                "auction_date_label": label,
                "date_precision": "year_range",
                "year_range": year_range,
                "count": len(rows),
                "file": shard_rel,
            }
        )
        auctions.append(
            {
                "auction_date": issue_key,
                "auction_date_label": label,
                "date_precision": "year_range",
                "year_range": year_range,
                "pdf_url": LEGACY_SOURCE_URL,
                "source_url": LEGACY_SOURCE_URL,
                "source_format": "xls",
                "source_type": "xls_legacy_year_range",
                "entry_count": len(rows),
                "total_sale_proceeds_hkd": total_amount,
                "error": None,
            }
        )

    preset_amount_desc = sorted(all_rows, key=compare_row)[:1000]
    manifest = {
        "total_rows": len(all_rows),
        "issue_count": len(manifest_items),
        "issues": manifest_items,
        "build_stats": stats,
    }

    (OUT / "results.slim.json").write_text(
        json.dumps(all_rows, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    (OUT / "auctions.json").write_text(
        json.dumps(list(reversed(auctions)), ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    (OUT / "issues.manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    (OUT / "preset.amount_desc.top1000.json").write_text(
        json.dumps(preset_amount_desc, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )

    print(
        json.dumps(
            {
                "dataset": "tvrm_legacy",
                "label": "1973-2006",
                "total_rows": len(all_rows),
                "issue_count": len(manifest_items),
                **stats,
            },
            ensure_ascii=False,
            separators=(",", ":"),
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(build())
