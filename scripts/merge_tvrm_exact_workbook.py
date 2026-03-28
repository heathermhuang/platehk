#!/usr/bin/env python3
from __future__ import annotations

import json
import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from urllib.parse import quote

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
EXACT_XLSX_NAME = "TVRM auction result (2006-2026).xlsx"
EXACT_XLSX_PATH = DATA / EXACT_XLSX_NAME
EXACT_SOURCE_URL = f"./data/{quote(EXACT_XLSX_NAME)}"
MERGE_START_YEAR = 2007
PLATE_RE = re.compile(r"^([A-Z]{1,2})(\d{1,4})$")


def normalize_plate(raw: str) -> str:
    value = re.sub(r"\s+", "", str(raw or "").upper())
    match = PLATE_RE.fullmatch(value)
    if match:
        return f"{match.group(1)} {match.group(2)}"
    return value


def normalize_amount(raw) -> int | None:
    text = str(raw or "").strip().replace(",", "")
    if text == "":
        return None
    try:
        return int(round(float(text)))
    except (TypeError, ValueError):
        return None


def normalize_date(raw) -> str | None:
    if isinstance(raw, datetime):
        return raw.date().isoformat()
    if isinstance(raw, date):
        return raw.isoformat()
    text = str(raw or "").strip()
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        return text
    return None


def result_status(text: str) -> str:
    text = str(text or "").strip()
    if "特別費用分配" in text:
        return "assigned_special_fee"
    if "售出" in text:
        return "sold"
    return "unknown"


def compare_row(row: dict) -> tuple:
    amount = row.get("amount_hkd")
    return (
        -(amount if amount is not None else -1),
        str(row.get("auction_date") or ""),
        str(row.get("single_line") or ""),
    )


def dedupe_key(row: dict) -> tuple:
    return (str(row.get("single_line") or ""), row.get("amount_hkd"))


def load_exact_rows_after_cutoff() -> dict[str, list[dict]]:
    by_issue: dict[str, list[dict]] = defaultdict(list)
    wb = load_workbook(EXACT_XLSX_PATH, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                plate_raw = row[0] if len(row) > 0 else None
                auction_raw = row[1] if len(row) > 1 else None
                amount_raw = row[2] if len(row) > 2 else None
                result_raw = row[3] if len(row) > 3 else None

                plate = normalize_plate(plate_raw)
                auction_date = normalize_date(auction_raw)
                if not plate or not auction_date or int(auction_date[:4]) < MERGE_START_YEAR:
                    continue
                amount = normalize_amount(amount_raw)
                result_text = str(result_raw or "").strip()
                status = result_status(result_text)
                by_issue[auction_date].append(
                    {
                        "auction_date": auction_date,
                        "single_line": plate,
                        "double_line": None,
                        "amount_hkd": amount,
                        "pdf_url": EXACT_SOURCE_URL,
                        "source_url": EXACT_SOURCE_URL,
                        "source_format": "xlsx",
                        "source_type": "xlsx_exact_dates",
                        "source_sheet": sheet_name,
                        "result_status": status,
                        "result_text": result_text,
                        "date_precision": "day",
                        "auction_date_label": auction_date,
                    }
                )
    finally:
        wb.close()
    return by_issue


def read_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, obj) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(obj, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")


def load_dataset_state(dataset_key: str) -> dict:
    base = DATA / dataset_key
    issues_dir = base / "issues"
    manifest = read_json(base / "issues.manifest.json")
    auctions = read_json(base / "auctions.json")
    auctions_by_date = {str(item["auction_date"]): dict(item) for item in auctions}
    rows_by_issue: dict[str, list[dict]] = {}
    for item in manifest.get("issues", []):
        issue_date = str(item["auction_date"])
        rows_by_issue[issue_date] = read_json(issues_dir / f"{issue_date}.json")
    return {
        "base": base,
        "dataset_key": dataset_key,
        "auctions_by_date": auctions_by_date,
        "rows_by_issue": rows_by_issue,
    }


def merge_issue_rows(existing_rows: list[dict], workbook_rows: list[dict], fallback_pdf_url: str) -> list[dict]:
    merged = {dedupe_key(row): dict(row) for row in existing_rows}
    for row in workbook_rows:
        key = dedupe_key(row)
        if key in merged:
            continue
        item = dict(row)
        item["pdf_url"] = fallback_pdf_url or item.get("pdf_url") or EXACT_SOURCE_URL
        merged[key] = item
    rows = list(merged.values())
    rows.sort(key=compare_row)
    return rows


def rebuild_dataset(dataset_key: str, state: dict) -> dict[str, int]:
    base: Path = state["base"]
    rows_by_issue: dict[str, list[dict]] = state["rows_by_issue"]
    auctions_by_date: dict[str, dict] = state["auctions_by_date"]
    issues_dir = base / "issues"
    issues_dir.mkdir(parents=True, exist_ok=True)

    issue_dates_asc = sorted(rows_by_issue.keys())
    issue_dates_desc = list(reversed(issue_dates_asc))
    all_rows: list[dict] = []
    auctions_out: list[dict] = []
    manifest_items: list[dict] = []

    keep_files = set()
    workbook_only_issues = 0

    for issue_date in issue_dates_asc:
        rows = rows_by_issue[issue_date]
        if not rows:
            continue
        keep_files.add(f"{issue_date}.json")
        rows.sort(key=compare_row)
        shard_rel = f"issues/{issue_date}.json"
        write_json(issues_dir / f"{issue_date}.json", rows)
        all_rows.extend(rows)

    for issue_date in issue_dates_desc:
        rows = rows_by_issue[issue_date]
        if not rows:
            continue
        manifest_items.append(
            {
                "auction_date": issue_date,
                "count": len(rows),
                "file": f"issues/{issue_date}.json",
            }
        )

    for issue_date in issue_dates_asc:
        rows = rows_by_issue[issue_date]
        if not rows:
            continue
        meta = dict(auctions_by_date.get(issue_date) or {})
        existing_pdf_url = str(meta.get("pdf_url") or "")
        if not meta:
            workbook_only_issues += 1
        total = sum(int(row["amount_hkd"]) for row in rows if row.get("amount_hkd") is not None)
        pdf_urls = meta.get("pdf_urls")
        if isinstance(pdf_urls, list) and pdf_urls:
            meta_pdf_urls = [str(x) for x in pdf_urls if str(x)]
        elif existing_pdf_url:
            meta_pdf_urls = [existing_pdf_url]
        else:
            meta_pdf_urls = [EXACT_SOURCE_URL]
        auction_label = meta.get("auction_date_label")
        if dataset_key == "tvrm_physical" and not auction_label:
            auction_label = ""
        elif not auction_label:
            auction_label = issue_date
        auctions_out.append(
            {
                "auction_date": issue_date,
                "auction_date_label": auction_label,
                "pdf_url": existing_pdf_url or EXACT_SOURCE_URL,
                "pdf_urls": meta_pdf_urls,
                "source_format": "xlsx" if any(str(x).lower().endswith(".xlsx") for x in meta_pdf_urls) else meta.get("source_format"),
                "source_type": "xlsx_exact_dates" if any(str(x).lower().endswith(".xlsx") for x in meta_pdf_urls) else meta.get("source_type"),
                "entry_count": len(rows),
                "total_sale_proceeds_hkd": total,
                "is_lny": bool(meta.get("is_lny")),
                "error": meta.get("error"),
            }
        )

    for path in issues_dir.glob("*.json"):
        if path.name not in keep_files:
            path.unlink()

    manifest = {
        "total_rows": len(all_rows),
        "issue_count": len(manifest_items),
        "issues": manifest_items,
    }
    preset_amount_desc = sorted(all_rows, key=compare_row)[:1000]

    write_json(base / "results.slim.json", all_rows)
    write_json(base / "issues.manifest.json", manifest)
    write_json(base / "auctions.json", auctions_out)
    write_json(base / "preset.amount_desc.top1000.json", preset_amount_desc)
    return {
        "rows": len(all_rows),
        "issues": len(manifest_items),
        "workbook_only_issues": workbook_only_issues,
    }


def merge() -> int:
    if not EXACT_XLSX_PATH.exists():
        raise FileNotFoundError(EXACT_XLSX_PATH)

    exact_by_issue = load_exact_rows_after_cutoff()
    physical = load_dataset_state("tvrm_physical")
    eauction = load_dataset_state("tvrm_eauction")
    eauction_dates = set(eauction["rows_by_issue"].keys())

    stats = {
        "merged_rows_added": 0,
        "issues_touched": 0,
        "physical_workbook_only_issues": 0,
        "eauction_workbook_only_issues": 0,
    }

    for issue_date, workbook_rows in exact_by_issue.items():
        target = eauction if issue_date in eauction_dates else physical
        issue_rows = target["rows_by_issue"].get(issue_date, [])
        fallback_pdf_url = str((target["auctions_by_date"].get(issue_date) or {}).get("pdf_url") or "")
        merged = merge_issue_rows(issue_rows, workbook_rows, fallback_pdf_url)
        added = len(merged) - len(issue_rows)
        if added <= 0:
            continue
        target["rows_by_issue"][issue_date] = merged
        stats["merged_rows_added"] += added
        stats["issues_touched"] += 1

    physical_stats = rebuild_dataset("tvrm_physical", physical)
    eauction_stats = rebuild_dataset("tvrm_eauction", eauction)
    stats["physical_workbook_only_issues"] = physical_stats["workbook_only_issues"]
    stats["eauction_workbook_only_issues"] = eauction_stats["workbook_only_issues"]

    print(json.dumps({"dataset": "tvrm_exact_merge", **stats}, ensure_ascii=False, separators=(",", ":")))
    return 0


if __name__ == "__main__":
    raise SystemExit(merge())
